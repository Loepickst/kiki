#!/usr/bin/env python3
"""Import N1 adverbs and loanwords from the master Numbers workbook.

The script treats the Numbers workbook as the content source of truth, while
keeping pitch accent handling conservative. It normalizes the accent values
found in the sheet, optionally applies explicit dictionary-verified overrides,
and writes a CSV audit table so uncertain items remain reviewable.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import tempfile
from collections import OrderedDict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - runtime guard
    raise SystemExit("openpyxl is required. Use the bundled workspace Python runtime.") from exc


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SOURCE = Path("/Users/kiki/Desktop/N1核心词汇.numbers")
N1_DIR = ROOT / "exam" / "vocabulary" / "n1"
AUDIT_PATH = N1_DIR / "n1_adverbs_loanwords_tone_audit.csv"

TONE_MARKS = {
    "⓪": "0",
    "①": "1",
    "②": "2",
    "③": "3",
    "④": "4",
    "⑤": "5",
    "⑥": "6",
    "⑦": "7",
    "⑧": "8",
    "⑨": "9",
}

# Keep this table intentionally small and auditable. Only add values that were
# explicitly verified against an external accent dictionary.
DICTIONARY_TONE_OVERRIDES: dict[str, dict[str, str]] = {
    "n1-adverbs": {},
    "n1-loanwords": {},
}


def normalize_space(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\u3000", " ")).strip()


def normalize_tone(value: Any) -> str:
    text = normalize_space(value)
    if not text:
        return ""
    for mark, digit in TONE_MARKS.items():
        text = text.replace(mark, digit)
    text = text.replace("／", "/").replace("・", "/")
    text = re.sub(r"[^0-9/]", "", text)
    text = re.sub(r"/+", "/", text).strip("/")
    return text


def extract_meaning_parts(text: str) -> list[dict[str, str]]:
    text = str(text or "").strip()
    if not text:
        return []
    parts: list[dict[str, str]] = []
    pattern = re.compile(r"[❶❷❸❹❺❻]\s*([^❶❷❸❹❺❻]+)")
    matches = pattern.findall(text)
    source_parts = matches if matches else [text]
    for raw in source_parts:
        raw = raw.strip()
        if not raw:
            continue
        m = re.match(r"([^（(]+)[（(](.*?)[）)]\s*$", raw)
        if m:
            title = normalize_space(m.group(1))
            detail = normalize_space(m.group(2))
        else:
            title = normalize_space(raw)
            detail = ""
        parts.append({"title": title, "detail": detail})
    return parts


def strip_number_prefix(text: str) -> str:
    return re.sub(r"^[❶❷❸❹❺❻]\s*[、.．]?\s*", "", str(text or "").strip())


def split_examples(text: str) -> list[dict[str, str]]:
    text = str(text or "").strip()
    if not text:
        return []
    chunks = re.split(r"(?=[❶❷❸❹❺❻]\s*[、.．]?\s*)", text)
    if len(chunks) <= 1:
        chunks = [text]
    examples: list[dict[str, str]] = []
    for chunk in chunks:
        chunk = strip_number_prefix(chunk)
        if not chunk:
            continue
        m = re.match(r"(.+?)（(.+?)）\s*$", chunk, flags=re.S)
        if m:
            jp = normalize_space(m.group(1))
            cn = normalize_space(m.group(2))
        else:
            jp = normalize_space(chunk)
            cn = ""
        examples.append({"jp": jp, "cn": cn})
    return examples


def highlight_word(sentence: str, word: str) -> str:
    if not sentence:
        return sentence
    escaped = re.escape(word)
    return re.sub(escaped, f"<span class='ex-highlight'>{word}</span>", sentence, count=1)


def make_word_html(word: str, reading: str | None = None) -> str:
    # These sheets already provide kana/katakana surface forms, so no ruby
    # synthesis is needed.
    return word


def build_common_fields(
    *,
    module_key: str,
    word: str,
    reading: str,
    tone: str,
    pos: str,
    meaning_text: str,
    collocation_text: str,
    examples_text: str,
    extra: dict[str, Any] | None = None,
) -> dict[str, Any]:
    meaning_parts = extract_meaning_parts(meaning_text)
    examples = split_examples(examples_text)
    if not meaning_parts:
        meaning_parts = [{"title": normalize_space(meaning_text), "detail": ""}]
    primary = meaning_parts[0]
    mean = "；".join(part["title"] for part in meaning_parts if part["title"])
    nuance_lines = []
    for index, part in enumerate(meaning_parts, start=1):
        if len(meaning_parts) > 1:
            prefix = f"{index}. "
        else:
            prefix = ""
        if part["detail"]:
            nuance_lines.append(f"{prefix}{part['title']}：{part['detail']}")
        else:
            nuance_lines.append(f"{prefix}{part['title']}")
    usage = f"{primary['title']}：{primary['detail']}" if primary.get("detail") else primary.get("title", "")

    collocations: list[dict[str, str]] = []
    if collocation_text:
        collocations.append({"jp": normalize_space(collocation_text), "cn": primary.get("title", "")})
    elif examples:
        collocations.append({"jp": examples[0]["jp"], "cn": examples[0].get("cn", "")})

    rendered_examples = []
    for example in examples:
        rendered_examples.append({
            "jp": highlight_word(example["jp"], word),
            "cn": example.get("cn", ""),
        })

    item: dict[str, Any] = {
        "word": word,
        "word_html": make_word_html(word, reading),
        "reading": reading,
        "tone": tone,
        "pos": pos,
        "mean": mean,
        "nuance": "\n".join(nuance_lines),
        "usage": usage,
        "collocation": collocations[0] if collocations else {"jp": "", "cn": ""},
        "collocations": collocations,
        "examples": rendered_examples,
    }
    if extra:
        item.update(extra)
    return item


def parse_loanword_meta(meta: str) -> tuple[str, str, str]:
    raw = normalize_space(meta)
    origin = ""
    origin_match = re.search(r"[（(](.*?)[）)]", raw)
    if origin_match:
        origin = normalize_space(origin_match.group(1))
    without_origin = re.sub(r"[（(].*?[）)]", "", raw).strip()
    pieces = [piece for piece in without_origin.split("・") if piece]
    tone = ""
    pos_parts: list[str] = []
    for piece in pieces:
        tone_candidate = normalize_tone(piece)
        if tone_candidate and re.fullmatch(r"[0-9/]+", tone_candidate):
            tone = tone_candidate
        else:
            pos_parts.append(piece)
    return "・".join(pos_parts) or "名", tone, origin


def choose_tone(module_key: str, word: str, source_tone: str) -> tuple[str, dict[str, str]]:
    override = DICTIONARY_TONE_OVERRIDES.get(module_key, {}).get(word)
    if override:
        status = "dictionary_confirmed" if override == source_tone else "dictionary_corrected"
        return override, {
            "dictionary_tone": override,
            "status": status,
            "source": "manual dictionary override",
            "note": "外部音调词典已确认。",
        }
    return source_tone, {
        "dictionary_tone": "",
        "status": "needs_manual_dictionary_check",
        "source": "Numbers normalized",
        "note": "自动访问 OJAD/Kanjium 失败，保留表格音调并列入人工校对。",
    }


def parse_adverbs(rows: list[tuple[Any, ...]], audit_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    merged: OrderedDict[str, dict[str, Any]] = OrderedDict()
    for row in rows:
        word = normalize_space(row[0])
        if not word:
            continue
        raw_tone = normalize_space(row[1])
        source_tone = normalize_tone(raw_tone)
        tone, audit = choose_tone("n1-adverbs", word, source_tone)
        entry = build_common_fields(
            module_key="n1-adverbs",
            word=word,
            reading=word,
            tone=tone,
            pos="副",
            meaning_text=str(row[2] or ""),
            collocation_text=str(row[3] or ""),
            examples_text=str(row[4] or ""),
        )
        if word in merged:
            target = merged[word]
            target["examples"].extend(example for example in entry["examples"] if example not in target["examples"])
            target["collocations"].extend(coll for coll in entry["collocations"] if coll not in target["collocations"])
            if target.get("collocations"):
                target["collocation"] = target["collocations"][0]
            target["nuance"] = "\n".join(dict.fromkeys([target["nuance"], entry["nuance"]]))
        else:
            merged[word] = entry
        audit_rows.append({
            "module": "n1-adverbs",
            "word": word,
            "source_tone_raw": raw_tone,
            "imported_tone": tone,
            **audit,
        })
    return list(merged.values())


def parse_loanwords(rows: list[tuple[Any, ...]], audit_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    items: list[dict[str, Any]] = []
    for row in rows:
        word = normalize_space(row[0])
        if not word:
            continue
        raw_meta = normalize_space(row[1])
        pos, source_tone, origin = parse_loanword_meta(raw_meta)
        tone, audit = choose_tone("n1-loanwords", word, source_tone)
        item = build_common_fields(
            module_key="n1-loanwords",
            word=word,
            reading=word,
            tone=tone,
            pos=pos,
            meaning_text=str(row[2] or ""),
            collocation_text="",
            examples_text=str(row[3] or ""),
            extra={"origin": origin},
        )
        items.append(item)
        audit_rows.append({
            "module": "n1-loanwords",
            "word": word,
            "source_tone_raw": raw_meta,
            "imported_tone": tone,
            **audit,
        })
    return items


def add_ids(items: list[dict[str, Any]], prefix: str) -> list[dict[str, Any]]:
    for index, item in enumerate(items, start=1):
        item["id"] = f"{prefix}-{index:03d}"
    return items


def js_string(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, indent=12)


def write_data_file(path: Path, items: list[dict[str, Any]]) -> None:
    text = "(function(global) {\n"
    text += "    'use strict';\n\n"
    text += "    const fullWordList = "
    text += json.dumps(items, ensure_ascii=False, indent=8)
    text += ";\n\n"
    text += "    global.FULL_WORD_LIST = fullWordList;\n"
    text += "    global.ALL_DEFINED_WORDS = fullWordList;\n"
    text += "    global.WORDS_PER_DAY = 20;\n"
    text += "})(typeof window !== 'undefined' ? window : globalThis);\n"
    path.write_text(text, encoding="utf-8")


def build_curriculum(items: list[dict[str, Any]], label: str, key: str) -> dict[str, Any]:
    days = []
    for start in range(0, len(items), 20):
        day_id = len(days) + 1
        word_ids = list(range(start, min(start + 20, len(items))))
        days.append({
            "id": day_id,
            "title": label,
            "subtitle": f"{len(word_ids)} 词",
            "groupId": 1,
            "groupTitle": label,
            "wordIds": word_ids,
        })
    stages = []
    for start in range(0, len(days), 2):
        stage_id = len(stages) + 1
        day_ids = [day["id"] for day in days[start:start + 2]]
        stages.append({
            "id": stage_id,
            "title": f"阶段 {stage_id} · {label}",
            "subtitle": f"单元 {day_ids[0]} - {day_ids[-1]}",
            "dayIds": day_ids,
            "groupTitle": label,
        })
    return {
        "groups": [{"id": 1, "title": label, "subtitle": f"{len(items)} 词"}],
        "days": days,
        "stages": stages,
        "key": key,
    }


def write_curriculum_file(path: Path, curriculum: dict[str, Any], globals_prefix: str) -> None:
    text = "(function(global) {\n"
    text += "    'use strict';\n\n"
    text += f"    const groups = {json.dumps(curriculum['groups'], ensure_ascii=False, indent=8)};\n"
    text += f"    const days = {json.dumps(curriculum['days'], ensure_ascii=False, indent=8)};\n"
    text += f"    const stages = {json.dumps(curriculum['stages'], ensure_ascii=False, indent=8)};\n\n"
    text += f"    global.{globals_prefix}_GROUPS = groups;\n"
    text += f"    global.{globals_prefix}_DAYS = days;\n"
    text += f"    global.{globals_prefix}_STAGES = stages;\n"
    text += "})(typeof window !== 'undefined' ? window : globalThis);\n"
    path.write_text(text, encoding="utf-8")


def write_audit(rows: list[dict[str, str]]) -> None:
    AUDIT_PATH.parent.mkdir(parents=True, exist_ok=True)
    fields = ["module", "word", "source_tone_raw", "imported_tone", "dictionary_tone", "status", "source", "note"]
    merged_rows: OrderedDict[tuple[str, str], dict[str, str]] = OrderedDict()
    for row in rows:
        key = (row["module"], row["word"])
        if key not in merged_rows:
            merged_rows[key] = dict(row)
            continue
        existing = merged_rows[key]
        raw_values = [value for value in existing["source_tone_raw"].split(" ; ") if value]
        if row["source_tone_raw"] not in raw_values:
            raw_values.append(row["source_tone_raw"])
        existing["source_tone_raw"] = " ; ".join(raw_values)
        existing["note"] = f"{existing['note']} 重复词条已合并。"
    with AUDIT_PATH.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(merged_rows.values())


def export_numbers_to_xlsx(source: Path) -> Path:
    soffice = shutil.which("soffice") or "/opt/homebrew/bin/soffice"
    # LibreOffice on macOS is much more reliable when exporting Numbers files
    # into /private/tmp than into a freshly-created nested temp directory.
    temp_dir = Path("/private/tmp")
    result = subprocess.run(
        [soffice, "--headless", "--convert-to", "xlsx", "--outdir", str(temp_dir), str(source)],
        capture_output=True,
        text=True,
        check=False,
    )
    if result.returncode != 0:
        fallback = temp_dir / f"{source.stem}.xlsx"
        if fallback.exists():
            print(f"Warning: Numbers export failed; using existing exported workbook at {fallback}")
            return fallback
        raise SystemExit(f"Failed to export Numbers workbook:\n{result.stdout}\n{result.stderr}")
    exported = temp_dir / f"{source.stem}.xlsx"
    if not exported.exists():
        matches = list(temp_dir.glob("*.xlsx"))
        if not matches:
            raise SystemExit(f"Numbers export did not produce an .xlsx file in {temp_dir}")
        exported = matches[0]
    return exported


def find_sheet(workbook, keyword: str):
    for sheet in workbook.worksheets:
        if keyword in sheet.title:
            return sheet
    raise SystemExit(f"Could not find sheet containing {keyword!r}. Available: {workbook.sheetnames}")


def read_rows(sheet) -> list[tuple[Any, ...]]:
    rows = list(sheet.iter_rows(values_only=True))
    return [row for row in rows[1:] if any(cell is not None and str(cell).strip() for cell in row)]


def main() -> None:
    parser = argparse.ArgumentParser(description="Import N1 adverbs and loanwords from Numbers.")
    parser.add_argument("--source", default=str(DEFAULT_SOURCE), help="Path to N1核心词汇.numbers or exported .xlsx")
    args = parser.parse_args()
    source = Path(args.source).expanduser()
    xlsx = source if source.suffix.lower() == ".xlsx" else export_numbers_to_xlsx(source)

    workbook = load_workbook(xlsx, read_only=True, data_only=True)
    adverb_rows = read_rows(find_sheet(workbook, "副词"))
    loanword_rows = read_rows(find_sheet(workbook, "外来"))
    audit_rows: list[dict[str, str]] = []

    adverbs = add_ids(parse_adverbs(adverb_rows, audit_rows), "n1-adverb")
    loanwords = add_ids(parse_loanwords(loanword_rows, audit_rows), "n1-loanword")

    write_data_file(N1_DIR / "n1_adverbs_data.js", adverbs)
    write_data_file(N1_DIR / "n1_loanwords_data.js", loanwords)
    write_curriculum_file(
        N1_DIR / "n1-adverbs-curriculum.js",
        build_curriculum(adverbs, "N1 副詞", "adverbs"),
        "LAB_N1_ADVERB",
    )
    write_curriculum_file(
        N1_DIR / "n1-loanwords-curriculum.js",
        build_curriculum(loanwords, "N1 外来語", "loanwords"),
        "LAB_N1_LOANWORD",
    )
    write_audit(audit_rows)

    print(f"Imported {len(adverbs)} adverbs and {len(loanwords)} loanwords.")
    print(f"Wrote tone audit: {AUDIT_PATH}")


if __name__ == "__main__":
    main()
